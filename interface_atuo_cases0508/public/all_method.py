# -*- coding: utf-8 -*-
# @Time    : 2020/10/31 10:30
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : all_method.py
# @Software: PyCharm


import requests
from conf import Allpath
from public.get_mysql_info import getMysqlInfo
from public.config import config
import redis
import time
import datetime
from public.logger import Log
from public.imageMaker import img_my
import random

logger = Log('all_method', Allpath.log_path)


# 获取运行环境配置
def project_env():
    p_env = config().read_config(Allpath.case_conf_path, 'ENV', 'project_env')
    if p_env == 'b2h':
        redis_db = config().read_config(Allpath.db_conf_path, 'REDIS', 'redis_b2h')
        db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_b2h')
        case_db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_case_db')
        http_url = 'B2HHTTP'
    elif p_env == 'h2b':
        redis_db = config().read_config(Allpath.db_conf_path, 'REDIS', 'redis_h2b')
        db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_h2b')
        case_db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_case_db')
        http_url = 'H2BHTTP'
    elif p_env == 'h2b2':
        redis_db = config().read_config(Allpath.db_conf_path, 'REDIS', 'redis_h2b2')
        db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_h2b2')
        case_db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_case_db')
        http_url = 'H2B2HTTP'
    elif p_env == 'h2b3':
        redis_db = config().read_config(Allpath.db_conf_path, 'REDIS', 'redis_h2b3')
        db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_h2b3')
        case_db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_case_db')
        http_url = 'H2B3HTTP'
    elif p_env == 'ZSHJ':
        redis_db = config().read_config(Allpath.db_conf_path, 'REDIS', 'redis_ZSHJ')
        db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_ZSHJ')
        case_db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_case_db')
        http_url = 'ZSHJHTTP'
    else:
        redis_db = ''
        db = ''
        case_db = config().read_config(Allpath.db_conf_path, 'DBCONFIG', 'config_case_db')
        http_url = ''
    return redis_db, db, case_db, http_url


projectenv = project_env()


# 在线报告地址获取
def ding_report_url():
    s = config().read_config(Allpath.case_conf_path, 'SHEET', 'sheet')
    if s == 'Sheet1':
        DD_name = "F端及微信单接口自动化测试报告"
    elif s == 'Sheet2':
        DD_name = "医院端个检团检逻辑接口自动化测试报告"
    elif s == 'Sheet3':
        DD_name = "企业直签逻辑接口自动化测试报告"
    elif s == 'Sheet4':
        DD_name = "腾讯健康接口自动化测试报告"
    elif s == 'Sheet5':
        DD_name = "阿里健康接口自动化测试报告"
    elif s == 'Sheet6':
        DD_name = "4.0重构个检团检逻辑接口自动化测试报告"
    elif s == 'Sheet7':
        DD_name = "微医接口逻辑接口自动化测试报告"
    elif s == 'Sheet8':
        DD_name = "正式环境GET接口自动化测试报告"
    else:
        DD_name = "｛请检查case表单配置｝接口自动化测试报告"
    return DD_name


# 钉钉消息机器人发送，及短信发送记录清理
def ding_msg(now):
    text = eval(
        getMysqlInfo(projectenv[2]).get_mysql_info_test("select * from jkgl ORDER BY date DESC LIMIT 10;", 1)[
            'restult'])
    logger.info('钉钉消息获取测试结果内容成功：%s' % text)
    bg_url = ding_report_url()
    s = config().read_config(Allpath.case_conf_path, 'SHEET', 'sheet')
    p_env = config().read_config(Allpath.case_conf_path, 'ENV', 'project_env')
    if p_env == 'ZSHJ':
        # 线上测试消息机器人
        url = 'https://oapi.dingtalk.com/robot/send?access_token=066c36699fd22b546f9560c8dbdf51ca9273c0e6c08ab11c73a4a1197ac2bfe9'

        # 个人测试地址
        #  url = 'https://oapi.dingtalk.com/robot/send?access_token=9abe37c6a640422cb85c9a877bb026b0b9029378890c5f14fcc106be9ab5725e'
    else:
        # 钉钉智能消息机器人测试
        url = 'https://oapi.dingtalk.com/robot/send?access_token=dfd67b41b306c26552aab304f54f98bbcfbaf31c39de2422b3b503aaea0f0a5d'

        # 个人测试地址
        # url = 'https://oapi.dingtalk.com/robot/send?access_token=9abe37c6a640422cb85c9a877bb026b0b9029378890c5f14fcc106be9ab5725e'
    headers = {"Content-Type": "application/json"}
    data = {"msgtype": "markdown", "markdown": {"title": bg_url,
                                                "text": "#### " + bg_url + "\n" + "> 【测试人员】:" + str(
                                                    text['testname']) + "\n" + "\n> 【开始时间】:" + str(
                                                    text['time']) + "\n" + "\n> 【合计耗时】:" + str(
                                                    text['sumtime']) + "\n" + "\n> 【本次结果】:" + str(
                                                    text['testresult']) + "\n" + "\n> 【通过率】:" + str(text['tonggl'])
                                                        + "\n" + "> ![screenshot](http://122.224.247.100:8500/"
                                                        + str(img_my()) + "_zgh.png)\n" + "> ###### " + now[0:10]
                                                        + " 发布 [在线查看报告详情](http://122.224.247.100:8508/" + s + ".html) \n"},
            "at": {"atMobiles": ["17681829051"], "isAtAll": "true"}}
    request = requests.post(url, json=data, headers=headers).json()
    logger.info('钉钉机器人消息请求内容{},钉钉机器人消息请求成功{}'.format(data, request))

    # 清理短信发送记录,初始化用户卡列表
    del_Husercard = getMysqlInfo(projectenv[1]).del_cardid_info()
    logger.info('清理短信发送记录成功！初始化用户卡列表成功！')


# F端登录获取公共header
def get_f_login_cookie():
    requests.packages.urllib3.disable_warnings()
    p_env = config().read_config(Allpath.case_conf_path, 'ENV', 'project_env')
    if p_env == 'ZSHJ':
        url = config().read_config(Allpath.http_conf_path, projectenv[3], 'url_f') + "/app/login/accountLogin"
        data = {'cellnumber': 15258814180, 'password': 'e10adc3949ba59abbe56e057f20f883e'}
        header = {'deviceId': '0A903CA5-C48E-46B9-86FE-E1562738F2AC', 'os': '1', 'version': '8.1.3',
                  'imei': '220DB5B0-7BBC-4AEC-8172-0C744F85C91F', 'from': '1', 'mac': 'f4:70:ab:60:1e:ea',
                  "openId": "o0Nj35mPANozjiNcdabTw9hWbAjw"}
    else:
        url = config().read_config(Allpath.http_conf_path, projectenv[3],
                                   'url_f') + "/healthmanage-web/app/login/accountLogin"
        data = {'cellnumber': 15258814180, 'password': '99fa72c8a55321a225c0a5abf0955585'}
        header = {'deviceId': 'Android-3806ba3c-be48-409e-b1d2-76d517348a64', 'os': '1', 'version': '8.1.3',
                  'imei': '864551031854596', 'from': '1', 'mac': 'f4:70:ab:60:1e:ea',
                  "openId": "oh02_51tYh6ybCZJgQ0LYtbb35bI"}
    # session = requests.Session()
    request = requests.post(url, data=data, headers=header, verify=False).json()
    header["token"] = request["result"]["token"]
    header["uid"] = request["result"]["userid"]
    logger.info("登录请求返回：{}===>公共header返回：{}".format(request, header))
    # cookie_jar = session.post(url1, data1, header).cookies
    # print(cookie_jar)
    # cookie = requests.utils.dict_from_cookiejar(cookie_jar)
    # logger.info("返回cookie：%s" % cookie)
    return header


# 医院后台根据手机号插入指定用户的token
def set_user_token(phone='15258814180', stationId='HL99998'):
    redis_db = projectenv[0]
    sql = {"my_sql": "select id from admin_user_common where  `account`= %s ", "condition": phone, "code": 0}
    sql_result = getMysqlInfo(projectenv[1]).get_mysql_info(sql['my_sql'], sql['condition'],
                                                            sql['code'])
    code = str(sql_result[0])
    logger.info('手机{}:id获取成功返回:{}'.format(phone, code))

    # 写入对应手机user_id的token进入redis
    # host是redis主机，需要redis服务端和客户端都起着 redis默认端口是6379
    pool = redis.ConnectionPool(host=redis_db['host'], password=redis_db['password'], port=redis_db['port'],
                                db=redis_db['db'], decode_responses=True)
    r = redis.Redis(connection_pool=pool)
    if phone == '15258814180':
        token = 'ee5c9f64a2569bb88f8aa23d38afc16f'
    else:
        token = 'ee999cd0bc8f5838fd28448a5f67fb14'
    # key是"gender" value是"male" 将键值对存入redis缓存
    r.set('healthmanage:comm:user:common:login:token:' + code, token)  # 医院后台

    headers = {'optimisticLockVersion': '', 'optimisticLockId': '', 'adminId': code, 'stationId': stationId,
               'token': token}
    logger.info('手机{}:header获取返回:{}'.format(phone, headers))
    return headers


# h2b/b2h运营后台根据手机号插入指定用户的token
def set_admin_token():
    redis_db = projectenv[0]
    # host是redis主机，需要redis服务端和客户端都起着 redis默认端口是6379
    pool = redis.ConnectionPool(host=redis_db['host'], password=redis_db['password'], port=redis_db['port'],
                                db=redis_db['db'], decode_responses=True)
    r = redis.Redis(connection_pool=pool)
    t = lock_time(0)

    if redis_db['db'] == 0:
        token = str(
            {"password": "e10adc3949ba59abbe56e057f20f883e", "modifyTime": t, "gender": 2, "createTime": t, "roleId": 1,
             "adminId": 121, "nickname": "测试账号", "userId": "17200000090", "account": "15258814180",
             "email": "11@qq.com", "status": 1, "token": "929f949bebd360b3892ad3942a6be7ff143"})
        r.set('healthmanage:web:AdminToken:143', token)  # h2b运营后台
        headers = {'uid': '17200000090', 'token': '929f949bebd360b3892ad3942a6be7ff143'}
    else:
        token = str(
            {"password": "e10adc3949ba59abbe56e057f20f883e", "modifyTime": t, "gender": 2, "createTime": t, "roleId": 1,
             "adminId": 121, "nickname": "测试账号", "userId": 17200000090, "account": "15258814180", "email": "11@qq.com",
             "status": 1, "token": "70fde9a9a2a19cf0718a690ba099c12c121"})
        r.set('healthmanage:web:AdminToken:121', token)  # b2h运营后台
        headers = {'uid': '17200000090', 'token': '70fde9a9a2a19cf0718a690ba099c12c121',
                   'god-portal-signature': 'c517aefa3e2ddcba789a60d47609369b540fd0d9502d1e212f6c6ea498b874b7',
                   'god-portal-timestamp': '1608261499685',
                   'god-portal-request-id': '6f02e800-3ebb-11eb-b067-e9a8515afd05'}
    # 线上：eede7c08f634931a37992466abfe0dbc470beffafb5c7fad936818fc086a49da 测试：c517aefa3e2ddcba789a60d47609369b540fd0d9502d1e212f6c6ea498b874b7
    logger.info('header获取返回:{}'.format(headers))
    return headers


# 获取图形验证码
def get_h_image_code(phone):
    # 转变为时间戳
    timeStamp = time.mktime(time.localtime())
    # print(int(timeStamp))
    # timeArray = time.localtime(1605257905)
    # otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    # print(otherStyleTime)
    url = config().read_config(Allpath.http_conf_path, projectenv[3],
                               'url_h') + "/healthmanage-hospital/admin/getValidateCode?account=" + phone + "&t=" + str(
        int(timeStamp))
    # data = {'account': 15258814180, 't': timeStamp}
    request = requests.get(url)
    logger.info("验证码请求返回：%s" % request)

    redis_db = projectenv[0]
    # host是redis主机，需要redis服务端和客户端都起着 redis默认端口是6379
    pool = redis.ConnectionPool(host=redis_db['host'], password=redis_db['password'], port=redis_db['port'],
                                db=redis_db['db'], decode_responses=True)
    r = redis.Redis(connection_pool=pool)
    # r.set('gender', 'male')  # key是"gender" value是"male" 将键值对存入redis缓存
    code = r.get('healthmanage:comm:user:common:login:code:image:' + phone)  # 医院端key，gender 取出键male对应的值
    # code = r.get('healthmanage: comm:adminvalidatecode_' + phone)  # 运营端的key
    logger.info("请求获取公共图形验证码返回：%s" % code)
    return code


# H端登录获取公共header
def get_h_login_cookie(phone='15258814180'):
    requests.packages.urllib3.disable_warnings()

    url = config().read_config(Allpath.http_conf_path, projectenv[3],
                               'url_h') + "/healthmanage-hospital/admin/checkValidateCode"
    data = {'account': phone, 'validateCode': get_h_image_code(phone)}
    sms_request = requests.post(url, params=data, verify=False).json()
    logger.info("图形验证码请求短信发送返回：%s" % sms_request)
    time.sleep(3)

    sql = {"my_sql": "select msg_content from msg_content where  `mobile_no`= %s ORDER BY id DESC", "condition": phone,
           "code": 0}
    sql_result = getMysqlInfo(projectenv[1]).get_mysql_info(sql['my_sql'], sql['condition'],
                                                            sql['code'])
    code = sql_result[0][6:11]
    logger.info('短信验证码获取成功返回:{}'.format(code))

    url = config().read_config(Allpath.http_conf_path, projectenv[3], 'url_h') + "/admin/user/login"
    data = {'account': phone, 'code': code, 'type': 1}
    header = {'adminId': '', 'stationId': '', 'token': ''}
    request = requests.post(url, params=data, headers=header, verify=False).json()
    logger.info("登录请求返回：%s" % request)
    header["token"] = request["result"]["token"]
    header["stationId"] = str(request["result"]["permissionDTOS"][0]['refId'])
    header["adminId"] = str(request["result"]["id"])

    logger.info("请求获取公共header返回：%s" % header)
    return header


# 获取微信端验证码
def message_code():
    requests.packages.urllib3.disable_warnings()
    p_env = config().read_config(Allpath.case_conf_path, 'ENV', 'project_env')
    if p_env == 'ZSHJ':
        url = config().read_config(Allpath.http_conf_path, projectenv[3], 'url_f') + "/wx/api/sendYzm"
    else:
        url = config().read_config(Allpath.http_conf_path, projectenv[3], 'url_f') + "/healthmanage-web/wx/api/sendYzm"
    data = {'cellphone': 15258814180}
    request = requests.get(url, data, verify=False).json()
    logger.info("短信请求返回：%s" % request)
    time.sleep(1)
    sql = {"my_sql": "select msg_content from msg_send_log where  `mobile_no`= %s ORDER BY id DESC",
           "condition": "15258814180", "code": 0}
    sql_result = getMysqlInfo(projectenv[1]).get_mysql_info(sql['my_sql'], sql['condition'],
                                                            sql['code'])
    text = sql_result[0][10:15]
    logger.info('短信验证码获取成功返回:{}'.format(text))
    return text


# 企业直签短信接口
def qyzq_msg_code():
    requests.packages.urllib3.disable_warnings()
    url = config().read_config(Allpath.http_conf_path, projectenv[3], 'url_f') + "/healthmanage-web/wx/api/smscode"
    data = {'cellphone': 15258814180, "valiImgCode": "", "isVerify": True}
    request = requests.get(url, data, verify=False).json()
    logger.info("短信请求返回：%s" % request)
    time.sleep(1)
    sql = {"my_sql": "select msg_content from msg_send_log where  `mobile_no`= %s ORDER BY id DESC",
           "condition": "15258814180", "code": 0}
    sql_result = getMysqlInfo(projectenv[1]).get_mysql_info(sql['my_sql'], sql['condition'],
                                                            sql['code'])
    text = sql_result[0][10:15]
    logger.info('短信验证码获取成功返回:{}'.format(text))
    return text


# 获取长效方案中时间戳校验值
def time_version(cxfa, phone='15258814180'):
    requests.packages.urllib3.disable_warnings()

    api = {'个检套餐': '/healthmanage-hospital/admin/suite/v2/newGoodsPkg/14617',
           '团检套餐': '/healthmanage-hospital/admin/suite/v2/newGoodsPkg/14618',
           '项目包': '/healthmanage-hospital/admin/optional/pkg/get/186',
           '获取机构': '/healthmanage-hospital/admin/admin/inst/detail?stationId=HL99998',
           '批次信息': '/healthmanage-hospital/admin/card/v3/batch/2695',
           '体检列表': '/healthmanage-hospital/admin/item/examHospital/list/HL99998/1/10',
           '对账后台': '/healthmanage-hospital/admin/behalf/detail?orderId=HjkBehalf20201116161217719'}
    url = config().read_config(Allpath.http_conf_path, projectenv[3], 'url_h')
    t_head = set_user_token(phone=phone)
    if cxfa == '体检列表':
        logger.info(cxfa + '详情请求连接:{}'.format(url + api[cxfa]))
        request = requests.post(url + api[cxfa], json={"name": "", "addFrom": 0, "isAdd": ''}, headers=t_head,
                                verify=False).json()
        timecode = request['result']['list'][0]['optimisticLockVersion']
        timeid = request['result']['list'][0]['optimisticLockId']

    else:
        logger.info(cxfa + '详情请求连接:{}'.format(url + api[cxfa]))
        request = requests.get(url + api[cxfa], headers=t_head, verify=False).json()
        timecode = request['result']['optimisticLockVersion']
        timeid = request['result']['optimisticLockId']

    logger.info(cxfa + '详情获取成功返回:{}'.format(request))
    logger.info('成功返回时间戳获取:{},时间戳id:{}'.format(timecode, timeid))
    t_head['optimisticLockVersion'] = str(timecode)
    t_head['optimisticLockId'] = str(timeid)

    return t_head


# 锁定2天返回时间戳
def lock_time(lock):
    # 今天日期
    today = datetime.date.today()
    # print(today)
    # 昨天时间
    yesterday = today - datetime.timedelta(days=1)
    # 明天时间
    tomorrow = today + datetime.timedelta(days=1)
    # 后天时间
    acquire = today + datetime.timedelta(days=2)

    if lock == 0:
        # 转为时间数组
        timeArray = time.strptime(str(today) + ' 21:00:00', "%Y-%m-%d %H:%M:%S")
        # print(timeArray)
        # 转为时间戳
        timeStamp = int(time.mktime(timeArray)) * 1000
        # print(timeStamp)  # 1381419600
        logger.info('今天时间戳：{},{}'.format(today, timeStamp))
        return timeStamp

    elif lock == 1:
        # 转为时间数组
        timeArray = time.strptime(str(tomorrow) + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        # print(timeArray)
        # 转为时间戳
        timeStamp = int(time.mktime(timeArray)) * 1000
        # print(timeStamp)  # 1381419600
        logger.info('明天时间戳：{},{}'.format(tomorrow, timeStamp))
        return timeStamp

    else:
        # 转为时间数组
        timeArray = time.strptime(str(acquire) + ' 21:00:00', "%Y-%m-%d %H:%M:%S")
        # print(timeArray)
        # 转为时间戳
        timeStamp = int(time.mktime(timeArray)) * 1000
        # print(timeStamp)  # 1381419600
        logger.info('后天时间戳：{},{}'.format(acquire, timeStamp))
        return timeStamp

# assertEqual断言内容返回
def assert_value_result(data, assert_value):
    value = []
    for i in assert_value:
        re = eval(i)
        value.append(re)
    return str(value)


# 根据用例id返回在这之前指定用例接口返回内容,通过指定路径value，重新合成data
def data_value_result(data, assert_value):
    try:
        result = eval(read_test_result(assert_value['id']))
    except:
        result = read_test_result(assert_value['id'])
    k = list(assert_value.keys())
    j = []
    for i in range(1, len(k)):
        re = eval(assert_value[k[i]])
        j.append(re)
    data = data % tuple(j)

    logger.info('调用id和返回字段{}\n获取指定返回值成功：{}'.format(assert_value, data))
    return data


# 根据用例idlist列表返回在这之前指定用例接口返回内容,通过指定路径value，重新合成data
def idlist_value_result(data, assert_value):
    id = assert_value['idlist']
    k = list(assert_value.keys())
    j = []
    for i in range(1, len(k)):
        try:
            result = eval(read_test_result(id[i - 1]))
        except:
            result = read_test_result(id[i - 1])
        re = eval(assert_value[k[i]])
        j.append(re)
    data = data % tuple(j)

    logger.info('调用id和返回字段{}\n获取指定返回值成功：{}'.format(assert_value, data))
    return data


# 根据用例id返回在这之前指定用例接口返回内容,通过指定路径value，重新合成url
def url_value_result(data, assert_value):
    try:
        result = eval(read_test_result(assert_value['path']))
    except:
        result = read_test_result(assert_value['path'])
    k = list(assert_value.keys())
    j = []
    for i in range(1, len(k)):
        re = eval(assert_value[k[i]])
        j.append(re)
    data = data % tuple(j)

    logger.info('调用id和返回字段{}\n获取指定返回值成功：{}'.format(assert_value, data))
    return data


# 根据sql查询作为入参
def sql_value_result(data, assert_value):
    sql_result = getMysqlInfo(projectenv[1]).get_mysql_info(assert_value['my_sql'],
                                                            assert_value['condition'], assert_value['code'])
    data = (data % tuple(sql_result))

    logger.info('调用my_sql进行查询和返回字段{}\n获取指定返回值成功：{}'.format(sql_result, data))
    return data


# 根据id获取单行接口请求返回结果值
def read_test_result(id):
    from openpyxl import load_workbook
    s = config().read_config(Allpath.case_conf_path, 'SHEET', 'sheet')
    file = Allpath.test_data_path
    rd = load_workbook(file)
    sheet = rd[s]
    value = str(sheet.cell(row=id + 1, column=13).value)
    logger.info('获取{}表单第{}条测试用例返回结果成功：{}'.format(s, id, value))
    return value


# 多字段不同用例返回结果取值方法，根据key为用例id，value为指定用例想获取值的路径
def many_data_result(data, assert_value):
    k = list(assert_value.keys())
    j = []
    for i in range(len(k)):
        try:
            result = eval(read_test_result(int(k[i])))
        except:
            result = read_test_result(int(k[i]))
        re = eval(assert_value[k[i]])
        j.append(re)
    data = data % tuple(j)

    logger.info('多字段不同用例返回调用id和返回字段{}\n获取指定返回值成功：{}'.format(assert_value, data))
    return data


# 随机名称获取替换
def data_value_replace(code, data):
    j = []
    for i in range(code):
        re = '接口测试' + str(datetime.datetime.now().__format__("%Y-%m-%d %H-%M-%S"))
        j.append(re)
    data = data % tuple(j)

    logger.info('调用随机名称获取替换字段返回值成功：{}'.format(data))
    return data


# 随机类型替换获取替换
def data_type_replace(code, data):
    j = []
    h = {}
    from faker import Faker
    faker = Faker(locale='zh_CN')
    if type(code) is list:
        for i in code:
            if i == "idcard":
                idcard = str(faker.ssn())
                j.append(idcard)
                h['idcard'] = idcard
            elif i == 'phone':
                phone = str(random.randint(10000000000, 13400000000))
                j.append(phone)
                h['phone'] = phone
    data = data % tuple(j)

    logger.info('调用随机类型获取替换字段返回值成功：{},{}'.format(data, h))
    return data, h


# 返回11位随机数替换url
def data_number_replace(code, url):
    j = []
    for i in range(code):
        re = str(random.randint(10000000000, 99999999999))
        j.append(re)
    url = url % tuple(j)

    logger.info('调用随机数获取替换字段返回值成功：{}'.format(url))
    return url, j


# 插入新流程单位到redis中
def mk_new_process(companyId):
    redis_db = projectenv[0]
    # host是redis主机，需要redis服务端和客户端都起着 redis默认端口是6379
    pool = redis.ConnectionPool(host=redis_db['host'], password=redis_db['password'], port=redis_db['port'],
                                db=redis_db['db'], decode_responses=True)
    r = redis.Redis(connection_pool=pool)

    code = r.get('healthmanage:userCard:companyStrategy')
    companyId = str(code) + ',' + str(companyId)
    r.set('healthmanage:userCard:companyStrategy', companyId)  # 插入新流程单位id
    logger.info('插入新流程单位id成功：{}'.format(companyId))


# 日期替换方法
def data_time_replace(data, env):
    if env['lock'] == 0:
        reserveDay = lock_time(0)
    else:
        reserveDay = lock_time(2)
    if "reserveDay" in data.keys():
        data['reserveDay'] = reserveDay
    elif "reserveDateTimeStamp" in data.keys():
        data['reserveDateTimeStamp'] = reserveDay
    else:
        a = eval(data['ext'])
        a['examinationReservePo']['reserveTime'] = reserveDay
        data['ext'] = a
    return data


# 初始化医院后台时间段可约
def order_time_init():
    requests.packages.urllib3.disable_warnings()

    url = config().read_config(Allpath.http_conf_path, projectenv[3],
                               'url_h') + "/healthmanage-hospital/admin/admin/inst/timeAppointment/modify"
    url1 = config().read_config(Allpath.http_conf_path, projectenv[3],
                                'url_h') + "/healthmanage-hospital/admin/admin/inst/modify"
    data = {"appointmentRatio": 100, "timeAppointmentFlag": 0, "stationId": "HL99998",
            "hospitalReserveIntervalConfigDTOList": []}
    data1 = {"instId": 1072, "reserveWeek": [1, 2, 3, 4, 5, 6, 7], "reserveRangeFrom": 0,
             "reserveForbiddenTime": "23:50:59", "drawBloodEndTime": "11:12", "drawBloodStartTime": "07:50",
             "reserveForbiddenDay": "", "whiteCompanyIds": "2708,2711,2734"}
    headers = set_user_token()
    request = requests.post(url, json=data, headers=headers, verify=False).json()
    request1 = requests.post(url1, json=data1, headers=headers, verify=False).json()
    logger.info("初始化后台时间段请求返回：{}===={}".format(request, request1))
    return request, request1


# 腾讯健康报告 Base64转PDF
def report_detail():
    import base64

    path = "../request_file/体检报告.pdf"
    with open("../request_file/base64.txt", "r") as f:  # 打开文件
        data = f.read()  # 读取文件

    print("读取base64：%s" % data)

    with open(path, 'wb') as f:
        f.write(base64.b64decode(data))
        print("体检报告解析成功！")


if __name__ == '__main__':
    '''2725,2768,2755'''
    get_f_login_cookie()
    # set_user_token()
    # message_code()
    # data = {"couponId": "%s", "bizChannel": 2, "userId": 17200000090, "stationId": "HL99997"}
    # value = {'result': [
    #     {'actionClient': '2', 'actionExtend': None, 'actionType': None, 'appUrl': '1', 'couponDeductType': 1,
    #      'couponDesc': None, 'couponId': 502, 'couponName': '测试优惠券0322', 'couponReceiveType': 1, 'couponRule': None,
    #      'couponStatus': 0, 'couponTotal': 100000, 'couponType': None, 'createTime': 1616401338000,
    #      'createTimeStr': '2021-03-22 16:22', 'deductMinValue': 0.0, 'deductValue': 3.0, 'goodsTypes': '个检券',
    #      'initiativeGetLink': 'http://management-fed-gray.helianhealth.com/envk8s//order/coupon-draw.html?couid=502',
    #      'isdelete': 0, 'jumpUrl': '1', 'modifyTime': 1616401338000, 'sendType': 3, 'showType': 1, 'spuIds': None,
    #      'takeCount': 3, 'usedCount': 1, 'validTime': '180120', 'validTimeStr': '2天2时2分', 'validType': 2,
    #      'wxUrl': '1'}], 'total': 476}
    # print(many_data_result(data,value))
    # e = projectenv
    # print(e)
    #
    # assert_value = {"329": "result['batchId']", "332": "result['normalData'][0]['bigBitchId']"}
    # code = {'testname': '测试组自动化报告', 'time': '2021-01-27 08:45:59', 'sumtime': '0:00:02.676459',
    #         'testresult': '共 281 条接口用例，通过 277 条，错误 4 条', 'tonggl': '98.58%'}
    # now = time.strftime('%Y-%m-%d_%H_%M_%S')
    #
    # # print(order_time_init())
    # ding_msg(now)
